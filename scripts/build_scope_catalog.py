#!/usr/bin/env python3
"""
Build the SAP S/4HANA Cloud Public Edition Scope Item Catalog.

Converts SAP's official "Availability and Dependencies" workbook into a
structured, graph-ready JSON reference held in mcp-server/catalog/scope_items.json.

This is a foundational knowledge asset for the whole agent fleet: every scope
item (e.g. J58, 1NT, BD9) carries its LOB, business area, component, provisioning,
dependency edges (required scope items), required master data, and country
availability. Retired scope items are captured separately as a do-not-use list.

Two consumers:
  1. MCP governance server  — structured lookup + dependency graph + retired guard.
  2. Public Cloud Brain     — one natural-language chunk per item for embeddings
                              (emit with --emit-chunks).

Source workbook (public SAP reference — not client data):
  Availability_Dependencies_EN_XX.xlsx  (Scope + Retired Scope Items sheets)

Usage:
    python scripts/build_scope_catalog.py <path-to-xlsx>
    python scripts/build_scope_catalog.py <path-to-xlsx> --emit-chunks

Install:
    pip install openpyxl
"""

import os
import re
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, timezone

BASE_DIR    = Path(__file__).resolve().parent.parent
CATALOG_OUT = BASE_DIR / "mcp-server" / "catalog" / "scope_items.json"
CHUNKS_OUT  = BASE_DIR / "brain" / "sap_reference" / "scope_items"

SCOPE_SHEET   = "Scope"
RETIRED_SHEET = "Retired Scope Items"

# Fixed column positions in the Scope sheet (header is row 2, data from row 3).
COL = {
    "lob": 0, "business_area": 1, "scope_item_id": 2, "description": 3,
    "cluster": 4, "component": 5, "additional_license": 6, "provisioning": 7,
    "selectable_in_scoping": 8, "more_info_optional": 9,
    "enterprise_management": 10, "public_sector": 11,
    "baseline_accelerator": 12, "retail": 13,
    "available_for_scope_extension": 14,
    "required_scope_items_raw": 15, "required_master_data_raw": 16,
}
COUNTRY_COL_START = 17   # columns 17..76 are ISO country codes (60 markets)

AUTHORITATIVE_SOURCE = (
    "https://help.sap.com/docs/SAP_S4HANA_CLOUD — Scope item catalog "
    "(Availability and Dependencies); confirm the current release in the tenant's "
    "SAP Central Business Configuration and the Best Practices Explorer."
)

# A scope-item / master-data ID is 3 alphanumerics (e.g. J58, 1NT, BNZ, 5XU).
_ID_RE = re.compile(r"[0-9A-Z]{3}")


def _clean(v):
    """Normalise a cell to a trimmed string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_required_scope_items(raw):
    """
    Parse the 'Required Scope Items and Business Conditions' cell into
    graph-ready edges. Notation:
      - semicolon-separated tokens
      - a plain token 'J45'  → hard-required dependency
      - a bracketed token '[BNZ]*' → conditional/optional dependency
    Returns (edges, raw). Each edge: {"to": <id>, "conditional": bool}.
    """
    if not raw:
        return [], None
    edges, seen = [], set()
    for token in str(raw).split(";"):
        token = token.strip()
        if not token:
            continue
        conditional = "[" in token or "*" in token
        m = _ID_RE.search(token)
        if not m:
            continue
        sid = m.group(0)
        if sid in seen:
            continue
        seen.add(sid)
        edges.append({"to": sid, "conditional": conditional})
    return edges, str(raw).strip()


def parse_master_data(raw):
    """
    Parse the 'Required Master Data' cell. Comma-separated 3-char IDs, possibly
    with country qualifiers like '(TR: ) BNP'. Returns (ids, raw).
    """
    if not raw:
        return [], None
    ids, seen = [], set()
    for token in str(raw).split(","):
        token = token.strip()
        if not token:
            continue
        # Strip any country qualifier prefix like "(TR: )"
        token = re.sub(r"\([^)]*\)", " ", token).strip()
        m = _ID_RE.search(token)
        if not m:
            continue
        mid = m.group(0)
        if mid not in seen:
            seen.add(mid)
            ids.append(mid)
    return ids, str(raw).strip()


def build(xlsx_path: Path, emit_chunks: bool):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── SCOPE ITEMS ───────────────────────────────────────────────────────────
    ws = wb[SCOPE_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = rows[0]
    country_codes = [
        _clean(header[i]) for i in range(COUNTRY_COL_START, len(header))
        if _clean(header[i])
    ]

    # A scope item is listed once per (LOB, Business Area) it touches, so the same
    # ID can span many rows with identical everything-else. Collapse to one record
    # per scope item, accumulating its classifications.
    by_id = {}          # scope_item_id -> record (insertion-ordered)
    lob_breakdown, prov_breakdown = {}, {}

    for row in rows[1:]:
        sid = _clean(row[COL["scope_item_id"]])
        if not sid:
            continue

        lob = _clean(row[COL["lob"]])
        ba  = _clean(row[COL["business_area"]])

        if sid not in by_id:
            req_edges, req_raw = parse_required_scope_items(row[COL["required_scope_items_raw"]])
            md_ids, md_raw     = parse_master_data(row[COL["required_master_data_raw"]])
            countries = [
                country_codes[j]
                for j, i in enumerate(range(COUNTRY_COL_START, len(header)))
                if j < len(country_codes)
                and _clean(row[i]) not in (None, "No")
            ]
            prov = _clean(row[COL["provisioning"]])
            prov_breakdown[prov] = prov_breakdown.get(prov, 0) + 1

            by_id[sid] = {
                "scope_item_id":   sid,
                "description":     _clean(row[COL["description"]]),
                "classifications": [],   # [{lob, business_area}] — filled below
                "lob":             lob,   # primary (first-seen) LOB
                "business_area":   ba,    # primary (first-seen) Business Area
                "cluster":         _clean(row[COL["cluster"]]),
                "component":       _clean(row[COL["component"]]),
                "provisioning":    prov,
                "additional_license_required":  _clean(row[COL["additional_license"]]),
                "selectable_in_scoping":         _clean(row[COL["selectable_in_scoping"]]),
                "available_for_scope_extension": _clean(row[COL["available_for_scope_extension"]]),
                "editions": {
                    "enterprise_management": _clean(row[COL["enterprise_management"]]),
                    "public_sector":         _clean(row[COL["public_sector"]]),
                    "baseline_accelerator":  _clean(row[COL["baseline_accelerator"]]),
                    "retail":                _clean(row[COL["retail"]]),
                },
                "required_scope_items":     req_edges,   # graph edges: [{to, conditional}]
                "required_scope_items_raw": req_raw,
                "required_master_data":     md_ids,
                "required_master_data_raw": md_raw,
                "available_countries":      countries,
                "available_country_count":  len(countries),
                "retired": False,
            }

        rec = by_id[sid]
        pair = {"lob": lob, "business_area": ba}
        if pair not in rec["classifications"]:
            rec["classifications"].append(pair)
        lob_breakdown[lob] = lob_breakdown.get(lob, 0) + 1   # counts every listing

    scope_items = list(by_id.values())

    # ── RETIRED SCOPE ITEMS ───────────────────────────────────────────────────
    retired = []
    rws = wb[RETIRED_SHEET]
    for row in rws.iter_rows(values_only=True):
        cell = _clean(row[0])
        if not cell or cell.lower().startswith("retired scope item"):
            continue
        # Format: "19C: Activity Management in Procurement"
        if ":" in cell:
            rid, desc = cell.split(":", 1)
            rid, desc = rid.strip(), desc.strip()
        else:
            rid, desc = cell.strip(), None
        if _ID_RE.fullmatch(rid):
            retired.append({"scope_item_id": rid, "description": desc, "retired": True})

    # ── ASSEMBLE ──────────────────────────────────────────────────────────────
    catalog = {
        "_meta": {
            "description": (
                "SAP S/4HANA Cloud Public Edition scope item catalog — the "
                "authoritative mapping of every scope item ID to its LOB, business "
                "area, component, provisioning, dependency graph, required master "
                "data, and country availability. Foundational reference for the "
                "agent fleet: resolves any scope-item reference (e.g. a BPD file "
                "prefix like 1MR/1NT) and links related items via the dependency graph."
            ),
            "authoritative_source": AUTHORITATIVE_SOURCE,
            "source_file": xlsx_path.name,
            "extracted_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "total_scope_items": len(scope_items),
            "total_listings": sum(len(s["classifications"]) for s in scope_items),
            "total_retired": len(retired),
            "country_markets": country_codes,
            "lob_breakdown": dict(sorted(lob_breakdown.items(), key=lambda x: -x[1])),
            "provisioning_breakdown": dict(sorted(prov_breakdown.items(), key=lambda x: -x[1])),
            "note": (
                "Public SAP product reference (not client data). Availability and "
                "dependencies change per release — re-run scripts/build_scope_catalog.py "
                "against the latest workbook to refresh. Always confirm the current "
                "state in the tenant's SAP Central Business Configuration."
            ),
        },
        "scope_items": scope_items,
        "retired_scope_items": retired,
    }

    CATALOG_OUT.parent.mkdir(parents=True, exist_ok=True)
    CATALOG_OUT.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] {CATALOG_OUT.relative_to(BASE_DIR)}")
    print(f"     {len(scope_items)} scope items, {len(retired)} retired, "
          f"{len(country_codes)} country markets")

    # ── OPTIONAL: EMBEDDINGS-READY CHUNKS (one per scope item) ─────────────────
    if emit_chunks:
        CHUNKS_OUT.mkdir(parents=True, exist_ok=True)
        for it in scope_items:
            deps = ", ".join(e["to"] for e in it["required_scope_items"]) or "none"
            md   = ", ".join(it["required_master_data"]) or "none"
            lobs = ", ".join(sorted({c["lob"] for c in it["classifications"] if c["lob"]}))
            bas  = ", ".join(sorted({c["business_area"] for c in it["classifications"] if c["business_area"]}))
            text = (
                f"SAP S/4HANA Cloud Public Edition scope item {it['scope_item_id']}: "
                f"{it['description']}. Lines of Business: {lobs}. "
                f"Business Areas: {bas}. Application component: {it['component']}. "
                f"Provisioning: {it['provisioning']}. "
                f"Required scope items (dependencies): {deps}. "
                f"Required master data: {md}. "
                f"Available in {it['available_country_count']} country markets."
            )
            (CHUNKS_OUT / f"{it['scope_item_id']}.json").write_text(
                json.dumps({
                    "id":               f"scope_{it['scope_item_id']}",
                    "source":           "SAP Scope Item Catalog",
                    "scope_item_id":    it["scope_item_id"],
                    "lob":              it["lob"],
                    "business_area":    it["business_area"],
                    "deliverable_type": "scope_item_reference",
                    "content_type":     "reference",
                    "text":             text,
                }, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[OK] {len(scope_items)} scope-item chunks -> "
              f"{CHUNKS_OUT.relative_to(BASE_DIR)}/")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", help="Path to Availability_Dependencies_EN_XX.xlsx")
    ap.add_argument("--emit-chunks", action="store_true",
                    help="Also emit one embeddings-ready chunk per scope item")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")
    build(xlsx_path, args.emit_chunks)


if __name__ == "__main__":
    main()
